#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
XLSX 读取工具。解析 4 行表头结构的 XLSX 文件，输出 JSON 格式数据。
用于验证 json_to_xlsx.py 生成结果的正确性。
"""

import json
import os
import sys
import argparse
from openpyxl import load_workbook


def read_xlsx(xlsx_path, sheet_name=None, max_rows=None):
    """
    读取 XLSX 文件并解析 4 行表头结构。
    返回 {headers: {...}, data: [...]} 字典。
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 4:
        return {'error': f'Expected at least 4 header rows, got {len(rows)}'}

    display_names = [str(v) if v is not None else '' for v in rows[0]]
    type_decls = [str(v) if v is not None else '' for v in rows[1]]
    field_names = [str(v) if v is not None else '' for v in rows[2]]
    format_hints = [str(v) if v is not None else '' for v in rows[3]]

    num_cols = len(field_names)
    while num_cols > 0 and not field_names[num_cols - 1]:
        num_cols -= 1

    display_names = display_names[:num_cols]
    type_decls = type_decls[:num_cols]
    field_names = field_names[:num_cols]
    format_hints = format_hints[:num_cols]

    headers = {
        'display_names': display_names,
        'type_declarations': type_decls,
        'field_names': field_names,
        'format_hints': format_hints
    }

    data_rows = rows[4:]
    if max_rows is not None:
        data_rows = data_rows[:max_rows]

    data = []
    for row in data_rows:
        record = {}
        row_vals = list(row) + [None] * max(0, num_cols - len(row))
        for i in range(num_cols):
            key = field_names[i]
            if not key:
                continue
            val = row_vals[i]
            record[key] = val
        if any(v is not None for v in record.values()):
            data.append(record)

    return {'headers': headers, 'data': data, 'total_rows': len(rows) - 4}


def main():
    parser = argparse.ArgumentParser(description='XLSX reader for 4-row header format')
    parser.add_argument('xlsx_path', help='Path to XLSX file')
    parser.add_argument('--sheet', '-s', default=None, help='Sheet name (default: active sheet)')
    parser.add_argument('--max-rows', '-n', type=int, default=None, help='Max data rows to read')
    parser.add_argument('--headers-only', action='store_true', help='Only print headers')
    parser.add_argument('--output', '-o', default=None, help='Output JSON file (default: stdout)')
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx_path):
        print(f"Error: file not found: {args.xlsx_path}")
        sys.exit(1)

    result = read_xlsx(args.xlsx_path, args.sheet, args.max_rows)

    if 'error' in result:
        print(f"Error: {result['error']}")
        sys.exit(1)

    if args.headers_only:
        output = result['headers']
        output['total_rows'] = result['total_rows']
    else:
        output = result

    json_str = json.dumps(output, ensure_ascii=False, indent=2, default=str)

    if args.output:
        with open(args.output, 'w', encoding='utf-8') as f:
            f.write(json_str)
        print(f"Written to {args.output}")
    else:
        print(json_str)


if __name__ == '__main__':
    main()
