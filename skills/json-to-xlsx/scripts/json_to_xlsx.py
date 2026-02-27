#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
JSON 配置文件批量转 XLSX 工具。
遵循 4 行表头结构：Row0=中文列名, Row1=类型声明, Row2=英文字段名, Row3=格式提示。
"""

import json
import os
import sys
import argparse
from collections import OrderedDict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


def flatten_object(obj, parent_key='', sep='.'):
    """将嵌套对象展开为扁平的 key-value 对。"""
    items = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else k
            if isinstance(v, dict):
                items.extend(flatten_object(v, new_key, sep).items())
            else:
                items.append((new_key, v))
    return OrderedDict(items)


def collect_all_keys(data_list):
    """从所有记录中收集完整的字段列表（保持出现顺序）。"""
    keys = OrderedDict()
    for record in data_list:
        flat = flatten_object(record)
        for k in flat:
            if k not in keys:
                keys[k] = None
    return list(keys.keys())


def infer_type_and_format(values):
    """
    根据一列所有非空值推断数据类型和格式提示。
    返回 (type_str, format_hint)。
    """
    non_none = [v for v in values if v is not None]
    if not non_none:
        return 'STRING', 'STRING'

    all_int = True
    all_float = True
    has_list = False
    list_samples = []

    for v in non_none:
        if isinstance(v, list):
            has_list = True
            list_samples.append(v)
            continue
        if isinstance(v, bool):
            all_int = False
            all_float = False
            continue
        if isinstance(v, int):
            continue
        if isinstance(v, float):
            all_int = False
            continue
        all_int = False
        all_float = False

    if has_list:
        fmt = infer_list_format(list_samples)
        return 'STRING', fmt

    if all_int and non_none:
        return 'INT', 'INT'
    if all_float and non_none:
        return 'FLOAT', 'FLOAT'
    return 'STRING', 'STRING'


def infer_list_format(list_samples):
    """
    根据数组样本推断格式提示。
    - 纯数字数组 -> [_]
    - 字符串数组含 '_' -> [_]
    - 字符串数组含 ',' -> [,]
    - 混合 -> [,_]
    """
    has_underscore = False
    has_comma = False
    for arr in list_samples:
        for item in arr:
            s = str(item)
            if '_' in s:
                has_underscore = True
            if ',' in s:
                has_comma = True

    if has_comma and has_underscore:
        return '[,_]'
    if has_comma:
        return '[,]'
    return '[_]'


def serialize_value(value):
    """将值序列化为 Excel 单元格内容。"""
    if value is None:
        return ''
    if isinstance(value, list):
        parts = [str(item) for item in value]
        return '_'.join(parts)
    if isinstance(value, bool):
        return 1 if value else 0
    if isinstance(value, (int, float)):
        return value
    return str(value)


def get_flat_value(record, key):
    """从原始记录中按展开后的 key 路径获取值。"""
    parts = key.split('.')
    current = record
    for part in parts:
        if isinstance(current, dict) and part in current:
            current = current[part]
        else:
            return None
    return current


def convert_json_to_xlsx(json_path, xlsx_path):
    """将单个 JSON 文件转换为 XLSX。"""
    with open(json_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    if not isinstance(data, list) or len(data) == 0:
        print(f"  [SKIP] {json_path}: not a non-empty array")
        return False

    keys = collect_all_keys(data)
    if not keys:
        print(f"  [SKIP] {json_path}: no fields found")
        return False

    all_values = {key: [] for key in keys}
    for record in data:
        for key in keys:
            val = get_flat_value(record, key)
            all_values[key].append(val)

    types_and_formats = {}
    for key in keys:
        types_and_formats[key] = infer_type_and_format(all_values[key])

    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    type_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, key in enumerate(keys, start=1):
        display_name = key.replace('.', '_')

        cell_r0 = ws.cell(row=1, column=col_idx, value=display_name)
        cell_r0.font = header_font
        cell_r0.fill = header_fill
        cell_r0.border = thin_border
        cell_r0.alignment = Alignment(horizontal='center')

        type_str, fmt_hint = types_and_formats[key]
        cell_r1 = ws.cell(row=2, column=col_idx, value=type_str)
        cell_r1.fill = type_fill
        cell_r1.border = thin_border
        cell_r1.alignment = Alignment(horizontal='center')

        cell_r2 = ws.cell(row=3, column=col_idx, value=key.replace('.', '_'))
        cell_r2.border = thin_border
        cell_r2.alignment = Alignment(horizontal='center')

        cell_r3 = ws.cell(row=4, column=col_idx, value=fmt_hint)
        cell_r3.fill = type_fill
        cell_r3.border = thin_border
        cell_r3.alignment = Alignment(horizontal='center')

    for row_idx, record in enumerate(data, start=5):
        for col_idx, key in enumerate(keys, start=1):
            raw_value = get_flat_value(record, key)
            cell_value = serialize_value(raw_value)
            cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = thin_border

    for col_idx, key in enumerate(keys, start=1):
        max_len = max(len(str(key)), 8)
        for row_idx in range(5, min(5 + len(data), 15)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None:
                max_len = max(max_len, min(len(str(cell_val)), 40))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    ws.freeze_panes = 'A5'

    wb.save(xlsx_path)
    wb.close()
    return True


def main():
    parser = argparse.ArgumentParser(description='JSON config to XLSX converter')
    parser.add_argument('input_dir', help='JSON 文件所在目录路径')
    parser.add_argument('--output-dir', '-o', default=None, help='XLSX output directory (default: <input-dir>/../xlsx/)')
    parser.add_argument('--files', '-f', nargs='*', help='Specific JSON filenames to convert (default: all)')
    args = parser.parse_args()

    input_dir = os.path.abspath(args.input_dir)
    if not os.path.isdir(input_dir):
        print(f"Error: input directory not found: {input_dir}")
        sys.exit(1)

    if args.output_dir:
        output_dir = os.path.abspath(args.output_dir)
    else:
        output_dir = os.path.join(os.path.dirname(input_dir), 'xlsx')

    os.makedirs(output_dir, exist_ok=True)

    json_files = []
    if args.files:
        for fname in args.files:
            if not fname.endswith('.json'):
                fname += '.json'
            fpath = os.path.join(input_dir, fname)
            if os.path.isfile(fpath):
                json_files.append(fname)
            else:
                print(f"Warning: file not found: {fpath}")
    else:
        json_files = sorted([f for f in os.listdir(input_dir) if f.endswith('.json')])

    if not json_files:
        print("No JSON files found.")
        sys.exit(1)

    print(f"Input:  {input_dir}")
    print(f"Output: {output_dir}")
    print(f"Files: {len(json_files)}")
    print()

    success = 0
    failed = 0
    for fname in json_files:
        json_path = os.path.join(input_dir, fname)
        xlsx_name = os.path.splitext(fname)[0] + '.xlsx'
        xlsx_path = os.path.join(output_dir, xlsx_name)
        try:
            if convert_json_to_xlsx(json_path, xlsx_path):
                print(f"  [OK] {fname} -> {xlsx_name}")
                success += 1
            else:
                failed += 1
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}")
            failed += 1

    print(f"\nDone: {success} success, {failed} failed")


if __name__ == '__main__':
    main()
